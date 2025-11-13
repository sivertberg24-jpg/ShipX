from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import List, Tuple, Optional
import re
from math import atan2, degrees, pi

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# =========================
# Data classes
# =========================

@dataclass
class Re1Header:
    cardids: List[str]
    rho_sw: float
    grav: float
    lpp: float
    breadth: float
    draught: float
    lcg: float
    vcg: float
    novel: int
    nohead: int
    nofreq: int
    ndof: int

@dataclass
class SpeedBlock:
    vel: float
    sink: float
    trim: float
    xmtn: Optional[float]
    zmtn: Optional[float]
    heads: List[float]
    freqs: List[float]
    # [nohead][nofreq][ndof] -> (real, imag)
    rao_re_im: List[List[List[Tuple[float, float]]]]

@dataclass
class Re1Data:
    header: Re1Header
    speeds: List[SpeedBlock]

# ---- NEW: phase-zeroing config to mimic the viewer ----
@dataclass
class PhaseZeroingConfig:
    enable: bool = True
    # Translational thresholds (m/m)
    trans_thresh: float = 1e-7          # Surge, Sway, Heave
    # Rotational thresholds (deg/m) -- since we export rotations in deg/m by default
    roll_thresh: float = 0.0            # Roll: leave as-is unless you want to mimic harder
    pitch_thresh: float = 0.0030        # Pitch
    yaw_thresh: float = 0.0040          # Yaw
    # If you see a viewer always forcing Surge phase to 0°, you can emulate that:
    force_surge_zero: bool = False

# =========================
# Parser (robust)
# =========================

# Accept both E/e and D/d exponents
_FLOAT = re.compile(r'[-+]?(?:\d+(?:\.\d*)?|\.\d+)(?:[EeDd][-+]?\d+)?')
_INT   = re.compile(r'[-+]?\d+')

def _normalize(line: str) -> str:
    # Normalize Fortran D exponents to E so float() always works
    return re.sub(r'([0-9.])[dD]([+-]?\d+)', r'\1E\2', line)

def _parse_floats(line: str) -> List[float]:
    s = _normalize(line)
    return [float(m.group(0)) for m in _FLOAT.finditer(s)]

def _parse_ints(line: str) -> List[int]:
    return [int(m.group(0)) for m in _INT.finditer(line)]

def _next_nonempty(lines: List[str], i: int) -> int:
    n = len(lines)
    while i < n and lines[i].strip() == "":
        i += 1
    return i

def parse_re1(path: str | Path) -> Re1Data:
    """
    Parse a VERES .re1 (Motion transfer functions) file.

    Robust to:
      * variable number of text 'CARDID' lines after the banner (not always 5),
      * tabs vs spaces,
      * Fortran D-exponents,
      * interleaved RAO blocks (FREQ line before each NDOF block) and 'freq-table-first' variant.
    """
    p = Path(path)
    with p.open('r', encoding='latin-1', errors='ignore') as f:
        # Keep original spacing (tabs/spaces) and blank lines
        lines = [ln.rstrip("\n") for ln in f]

    idx = 0
    n = len(lines)

    def cur() -> str:
        return lines[idx]

    # 1) Find the banner
    while idx < n and "MOTION TRANSFER FUNCTIONS" not in cur().upper():
        idx += 1
    if idx >= n:
        raise ValueError("Could not find 'MOTION TRANSFER FUNCTIONS' banner in file.")
    idx += 1  # move past banner

    # 2) Collect CARDIDs until we can positively identify the numeric header start.
    # Detect ρSW/GRAV as: a line with ≥2 floats whose NEXT non-empty line has ≥3 floats (LPP/B/T).
    cardids: List[str] = []
    while idx < n:
        idx = _next_nonempty(lines, idx)
        if idx >= n:
            break
        s = lines[idx]
        f_here = _parse_floats(s)
        if len(f_here) >= 2:
            j = _next_nonempty(lines, idx + 1)
            if j < n and len(_parse_floats(lines[j])) >= 3:
                # This is the ρSW/GRAV line
                break
        # Otherwise, treat as a CARDID/text line
        cardids.append(s.strip())
        idx += 1
    if idx >= n:
        raise ValueError("Unexpected end of file while searching for RHOSW/GRAV line.")

    # 3) RHOSW, GRAV
    vals = _parse_floats(cur()); idx += 1
    if len(vals) < 2:
        raise ValueError("Expected two floats for RHOSW GRAV.")
    rho_sw, grav = vals[0], vals[1]

    # 4) LPP, BREADTH, DRAUGHT
    idx = _next_nonempty(lines, idx)
    if idx >= n:
        raise ValueError("Unexpected end of file before LPP/BREADTH/DRAUGHT.")
    vals = _parse_floats(cur()); idx += 1
    if len(vals) < 3:
        raise ValueError("Expected three floats for LPP BREADTH DRAUGHT.")
    lpp, breadth, draught = vals[:3]

    # 5) LCG, VCG
    idx = _next_nonempty(lines, idx)
    vals = _parse_floats(cur()); idx += 1
    if len(vals) < 2:
        raise ValueError("Expected two floats for LCG VCG.")
    lcg, vcg = vals[0], vals[1]

    # 6) NOVEL, NOHEAD, NOFREQ, [NDOF]
    idx = _next_nonempty(lines, idx)
    ints = _parse_ints(cur()); idx += 1
    if len(ints) == 4:
        novel, nohead, nofreq, ndof = ints
    elif len(ints) == 3:
        novel, nohead, nofreq = ints
        ndof = 6
    else:
        raise ValueError("Expected 3 or 4 integers for NOVEL NOHEAD NOFREQ [NDOF].")

    header = Re1Header(
        cardids=cardids, rho_sw=rho_sw, grav=grav,
        lpp=lpp, breadth=breadth, draught=draught,
        lcg=lcg, vcg=vcg, novel=novel, nohead=nohead, nofreq=nofreq, ndof=ndof
    )

    speeds: List[SpeedBlock] = []

    # 7) For each speed block
    for _ivel in range(novel):
        # VEL SINK TRIM [XMTN ZMTN]
        idx = _next_nonempty(lines, idx)
        if idx >= n:
            raise ValueError("Unexpected end of file before VEL/SINK/TRIM.")
        vals = _parse_floats(cur()); idx += 1
        if len(vals) < 3:
            raise ValueError("Expected at least three floats for VEL SINK TRIM.")
        vel, sink, trim = vals[0], vals[1], vals[2]
        xmtn = vals[3] if len(vals) >= 4 else None
        zmtn = vals[4] if len(vals) >= 5 else None

        # HEAD list (nohead lines)
        heads: List[float] = []
        for _ in range(nohead):
            idx = _next_nonempty(lines, idx)
            if idx >= n:
                raise ValueError("Unexpected end of file while reading HEAD list.")
            hvals = _parse_floats(cur())
            if not hvals:
                raise ValueError(f"HEAD line contained no float: '{cur()}'")
            heads.append(hvals[-1])  # take last number on the line
            idx += 1

        # ---- RAO + Frequencies ----
        idx_after_heads = idx
        try:
            idx, freqs, rao_re_im = _parse_speed_block_interleaved(
                lines, idx_after_heads, nohead, nofreq, ndof
            )
        except Exception:
            # Fallback: explicit freq table first, then RAOs
            idx, freqs, rao_re_im = _parse_speed_block_with_freq_table(
                lines, idx_after_heads, nohead, nofreq, ndof
            )

        speeds.append(SpeedBlock(
            vel=vel, sink=sink, trim=trim, xmtn=xmtn, zmtn=zmtn,
            heads=heads, freqs=freqs, rao_re_im=rao_re_im
        ))

    return Re1Data(header=header, speeds=speeds)

def _parse_speed_block_interleaved(
    lines: List[str], idx0: int, nohead: int, nofreq: int, ndof: int
) -> Tuple[int, List[float], List[List[List[Tuple[float, float]]]]]:
    """
    Interleaved: for each head & freq
      FREQ line (just the value or contains 'FREQ') -> then NDOF RAO lines.
    """
    idx = idx0
    freqs: List[float] = []
    rao_re_im: List[List[List[Tuple[float, float]]]] = []

    def parse_freq_line(s: str) -> Optional[float]:
        f = _parse_floats(s)
        up = s.upper()
        if "FREQ" in up and f:
            return f[-1]
        if len(f) == 1:
            return f[0]
        return None

    for ih in range(nohead):
        rao_re_im.append([])
        for ifr in range(nofreq):
            idx = _next_nonempty(lines, idx)
            if idx >= len(lines):
                raise ValueError("Unexpected end of file in interleaved FREQ/RAO section.")
            fcur = parse_freq_line(lines[idx])
            if fcur is None:
                raise ValueError(f"Expected FREQ line, got: '{lines[idx]}'")
            if ih == 0:
                freqs.append(fcur)
            else:
                if abs(fcur - freqs[ifr]) > 1e-9:
                    raise ValueError(f"Inconsistent frequency at head {ih}, freq {ifr}: {fcur} vs {freqs[ifr]}")
            idx += 1

            dof_vals: List[Tuple[float, float]] = []
            for _ldof in range(ndof):
                idx = _next_nonempty(lines, idx)
                if idx >= len(lines):
                    raise ValueError("Unexpected end of file while reading RAO line.")
                vals = _parse_floats(lines[idx])
                if len(vals) < 2:
                    raise ValueError(f"Could not parse real/imag on RAO line: '{lines[idx]}'")
                # Use last two floats to be tolerant of an optional leading DOF index
                re_val, im_val = vals[-2], vals[-1]
                dof_vals.append((re_val, im_val))
                idx += 1
            rao_re_im[ih].append(dof_vals)

    return idx, freqs, rao_re_im

def _parse_speed_block_with_freq_table(
    lines: List[str], idx0: int, nohead: int, nofreq: int, ndof: int
) -> Tuple[int, List[float], List[List[List[Tuple[float, float]]]]]:
    """
    Variant where a block of 'nofreq' frequency lines appears once before RAOs.
    """
    idx = idx0
    freqs: List[float] = []
    for _ in range(nofreq):
        idx = _next_nonempty(lines, idx)
        if idx >= len(lines):
            raise ValueError("Unexpected end of file while reading FREQLIST.")
        fvals = _parse_floats(lines[idx])
        if not fvals:
            raise ValueError(f"Frequency line contained no number: '{lines[idx]}'")
        freqs.append(fvals[-1])
        idx += 1

    rao_re_im: List[List[List[Tuple[float, float]]]] = []
    for ih in range(nohead):
        rao_re_im.append([])
        for _ifr in range(nofreq):
            dof_vals: List[Tuple[float, float]] = []
            for _ldof in range(ndof):
                idx = _next_nonempty(lines, idx)
                if idx >= len(lines):
                    raise ValueError("Unexpected end of file while reading RAO block.")
                vals = _parse_floats(lines[idx])
                if len(vals) < 2:
                    raise ValueError(f"Could not parse real/imag on RAO line: '{lines[idx]}'")
                re_val, im_val = vals[-2], vals[-1]
                dof_vals.append((re_val, im_val))
                idx += 1
            rao_re_im[ih].append(dof_vals)

    return idx, freqs, rao_re_im

# =========================
# Excel writer
# =========================

DOF_NAMES_DEFAULT = ["Surge", "Sway", "Heave", "Roll", "Pitch", "Yaw"]

def _knots(ms: float) -> float:
    return ms * 1.9438444924406

def _format_pos_ap(x_from_ap: float) -> str:
    side = "in front of AP" if x_from_ap >= 0 else "aft of AP"
    return f"{abs(x_from_ap):.3f} m {side}"

def _round_str(x: Optional[float]) -> str:
    return "0.000" if x is None else f"{x:.3f}"

def _apply_styles(ws, data_start_row: int, data_end_row: int):
    ws["A1"].font = Font(bold=True)
    for r in [25, 26]:
        for c in range(1, 14):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 26
    for col in range(2, 14):
        ws.column_dimensions[get_column_letter(col)].width = 13
    for r in range(1, 24):
        ws.row_dimensions[r].height = 18
    for r in range(data_start_row, data_end_row + 1):
        ws.cell(row=r, column=1).number_format = "0.000000"
        for col in range(2, 14):
            ws.cell(row=r, column=col).number_format = "0.000000"

# ---- NEW: helper to mimic the viewer's phase zeroing ----
def _maybe_zero_phase(
    dof_index: int,
    dof_name: str,
    amp_value: Optional[float],
    phase_value: Optional[float],
    zero_cfg: Optional[PhaseZeroingConfig],
    is_rotational_amp_in_deg_per_m: bool
) -> Optional[float]:
    if phase_value is None or amp_value is None or zero_cfg is None or not zero_cfg.enable:
        return phase_value

    # Translational DOFs: 0,1,2  (amp in m/m)
    if dof_index == 0:  # Surge
        if zero_cfg.force_surge_zero or abs(amp_value) < zero_cfg.trans_thresh:
            return 0.0
    if dof_index in (1, 2):  # Sway, Heave
        if abs(amp_value) < zero_cfg.trans_thresh:
            return 0.0

    # Rotational DOFs: 3,4,5
    if dof_index >= 3:
        # Compare in the *same units as amp_value*. Our default export uses deg/m.
        if is_rotational_amp_in_deg_per_m:
            roll_thr  = zero_cfg.roll_thresh
            pitch_thr = zero_cfg.pitch_thresh
            yaw_thr   = zero_cfg.yaw_thresh
        else:
            # If someday you export rotations in rad/m, convert deg/m thresholds to rad/m here:
            deg2rad = pi / 180.0
            roll_thr  = zero_cfg.roll_thresh  * deg2rad
            pitch_thr = zero_cfg.pitch_thresh * deg2rad
            yaw_thr   = zero_cfg.yaw_thresh   * deg2rad

        if dof_index == 3 and abs(amp_value) < roll_thr:
            return 0.0
        if dof_index == 4 and abs(amp_value) < pitch_thr:
            return 0.0
        if dof_index == 5 and abs(amp_value) < yaw_thr:
            return 0.0

    return phase_value

def _write_sheet(
    wb: Workbook,
    re1: Re1Data,
    speed: SpeedBlock,
    head_idx: int,
    dof_names: List[str] = DOF_NAMES_DEFAULT,
    sheet_name: Optional[str] = None,
    convert_rot_to_deg_per_m: bool = True,
    phase_zeroing: Optional[PhaseZeroingConfig] = None,  # NEW
):
    heads = speed.heads
    freqs = speed.freqs
    ndof  = re1.header.ndof
    nofreq = len(freqs)

    if head_idx < 0 or head_idx >= len(heads):
        raise IndexError("head_idx out of bounds")

    # Unique sheet name
    head_val = heads[head_idx]
    if sheet_name is None:
        base = f"H{int(head_val) if float(head_val).is_integer() else head_val:g}"
        name = base
        i = 1
        while name in wb.sheetnames:
            i += 1
            name = f"{base}_{i}"
        sheet_name = name[:31]
    else:
        sheet_name = sheet_name[:31]

    ws = wb.create_sheet(title=sheet_name)

    # Header text
    run_name = " ".join(re1.header.cardids).strip() or Path('.').resolve().name
    ws["A1"] = f"Transfer functions (RAOs) for Run name:  {run_name}"
    ws["A3"] = f"LPP = {re1.header.lpp:.3f} m, B = {re1.header.breadth:.3f} m"
    x_ap_cog = re1.header.lpp/2.0 + re1.header.lcg
    ws["A4"] = f"COG at {_format_pos_ap(x_ap_cog)}, 0.000 m from centerline, {_round_str(re1.header.vcg)} m from baseline"

    if speed.xmtn is not None and speed.zmtn is not None:
        x_ap_mtn = re1.header.lpp/2.0 + speed.xmtn
        ws["A5"] = f"MTN at {_format_pos_ap(x_ap_mtn)}, 0.000 m from centerline, {_round_str(speed.zmtn)} m from baseline"
    else:
        ws["A5"] = "MTN position not available in this .re1 file (old format; no XMTN/ZMTN)."

    ws["A6"] = f"Draught {re1.header.draught:.3f} m"
    ws["A8"] = "RAOs are presented for COG position"
    ws["A10"] = "Data in VERES format conventions:"
    ws["A11"] = "Positive SURGE is aft, SWAY is to starboard, HEAVE is up."
    ws["A12"] = "Positive ROLL is port down, PITCH is bow up and YAW is bow to starboard side."
    ws["A13"] = "PHASE ANGLES are positive in degrees lead to wave amplitude."
    ws["A14"] = "ROTATIONS are divided by wave amplitude (rad/m); amplitudes shown here are converted to deg/m."
    ws["A20"] = "270 degrees are waves from starboard (starboard bow to aft, positive clockwise heading)."

    ws["A22"] = "Vessel speed"; ws["B22"] = f"{_knots(speed.vel):.2f}"; ws["C22"] = "[knots]"
    ws["A23"] = "Heading";      ws["B23"] = f"{heads[head_idx]:.1f}";   ws["C23"] = "[deg] off bow"

    # Table headers
    header_row1, header_row2 = 25, 26
    ws.cell(row=header_row1, column=1, value="Wave frequency")
    ws.cell(row=header_row2, column=1, value="rad/sec")

    start_col = 2
    for i in range(6):
        group = dof_names[i] if i < len(dof_names) else f"DOF{i+1}"
        ws.cell(row=header_row1, column=start_col + 2*i,     value=group)
        ws.cell(row=header_row2, column=start_col + 2*i,     value="RAO (m/m)" if i < 3 else "RAO (deg/m)")
        ws.cell(row=header_row2, column=start_col + 2*i + 1, value="Phase (deg)")
        ws.merge_cells(start_row=header_row1, start_column=start_col + 2*i,
                       end_row=header_row1,   end_column=start_col + 2*i + 1)

    # Data rows
    data_start_row = header_row2 + 2
    for r in range(nofreq):
        row = data_start_row + r
        ws.cell(row=row, column=1, value=freqs[r])
        for i in range(6):
            if i < ndof:
                re_val, im_val = speed.rao_re_im[head_idx][r][i]
                amp = (re_val*re_val + im_val*im_val) ** 0.5
                phase = degrees(atan2(im_val, re_val))
                # Normalize to (-180, 180]
                phase = ((phase + 180.0) % 360.0) - 180.0
                # Convert rotations to deg/m if requested
                if i >= 3 and convert_rot_to_deg_per_m:
                    amp *= (180.0 / pi)  # rad/m -> deg/m
                # ---- NEW: mimic viewer by zeroing phases in low-amplitude regions ----
                dof_name = (dof_names[i] if i < len(dof_names) else f"DOF{i+1}")
                phase = _maybe_zero_phase(
                    dof_index=i,
                    dof_name=dof_name,
                    amp_value=amp,
                    phase_value=phase,
                    zero_cfg=phase_zeroing,
                    is_rotational_amp_in_deg_per_m=convert_rot_to_deg_per_m
                )
            else:
                amp, phase = None, None
            ws.cell(row=row, column=start_col + 2*i,     value=amp)
            ws.cell(row=row, column=start_col + 2*i + 1, value=phase)

    _apply_styles(ws, data_start_row, data_start_row + nofreq - 1)

# =========================
# Public API
# =========================

def convert_re1_to_excel(
    re1_path: str | Path,
    out_xlsx: Optional[str | Path] = None,
    speed_index: int = 0,
    heading_indices: Optional[List[int]] = None,
    sheet_name_prefix: Optional[str] = None,
    *,
    match_viewer_phase_zeroing: bool = True,   # NEW: on by default
    phase_zeroing_config: Optional[PhaseZeroingConfig] = None,  # NEW: override thresholds
) -> Path:
    """
    Convert a VERES .re1 file to an Excel workbook with one sheet per selected heading.

    Parameters
    ----------
    match_viewer_phase_zeroing : bool
        If True (default), zero phases for very small amplitudes to mimic the VERES viewer's output.
        Set to False to keep the mathematically exact atan2 phase everywhere.
    phase_zeroing_config : PhaseZeroingConfig | None
        Optional thresholds and flags; if None, sensible defaults are used.
    """
    re1 = parse_re1(re1_path)

    if speed_index < 0 or speed_index >= len(re1.speeds):
        raise IndexError(f"speed_index {speed_index} out of bounds (0..{len(re1.speeds)-1})")

    speed = re1.speeds[speed_index]
    if heading_indices is None:
        heading_indices = list(range(len(speed.heads)))

    # Prepare zeroing config
    zero_cfg = None
    if match_viewer_phase_zeroing:
        zero_cfg = phase_zeroing_config or PhaseZeroingConfig()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop default sheet

    for hi in heading_indices:
        sheet_name = None
        if sheet_name_prefix is not None:
            head_val = speed.heads[hi]
            base = f"{sheet_name_prefix}_H{int(head_val) if float(head_val).is_integer() else head_val:g}"
            sheet_name = base[:31]
        _write_sheet(
            wb, re1, speed, hi,
            sheet_name=sheet_name,
            convert_rot_to_deg_per_m=True,
            phase_zeroing=zero_cfg
        )

    out = Path(out_xlsx) if out_xlsx else Path(re1_path).with_suffix('').with_name(Path(re1_path).stem + "_rao.xlsx")
    wb.save(out)
    return out
